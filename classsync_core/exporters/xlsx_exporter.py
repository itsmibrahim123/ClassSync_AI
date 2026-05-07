"""
XLSX (Excel) exporter with styling and formatting.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, Optional
import os

from classsync_core.exports import BaseExporter


from classsync_core.models import ConstraintConfig, Room

class XLSXExporter(BaseExporter):
    """Export timetables to Excel format with styling."""

    def __init__(self, db):
        super().__init__(db)

        # Default styling
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.day_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export(self, timetable_id: int, output_path: str, **kwargs) -> str:
        """
        Export timetable to Excel file.

        Args:
            timetable_id: ID of timetable to export
            output_path: Path where file should be saved
            **kwargs: Options like 'view_type' (section/teacher/room/master/program/free_slots)

        Returns:
            Path to exported file
        """
        view_type = kwargs.get('view_type', 'master')

        # Load data
        df = self.load_timetable_data(timetable_id)

        if df.empty and view_type != 'free_slots':
             # Even for free_slots, we might need df to know what's occupied, but we definitely need it for others
             pass

        # Create output directory if needed
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

        if view_type == 'master':
            return self._export_master(df, output_path)
        elif view_type == 'section':
            return self._export_by_section(df, output_path)
        elif view_type == 'teacher':
            return self._export_by_teacher(df, output_path)
        elif view_type == 'room':
            return self._export_by_room(df, output_path)
        elif view_type == 'program':
            return self._export_by_program(df, output_path)
        elif view_type == 'free_slots':
            return self._export_free_slots(df, output_path, timetable_id)
        else:
            raise ValueError(f"Unknown view type: {view_type}")

    def _export_master(self, df: pd.DataFrame, output_path: str) -> str:
        """Export complete timetable as single Excel file."""

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Timetable"

        # Group by day for better visualization
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

        row = 1
        for day in days:
            day_df = df[df['Weekday'] == day].sort_values('Start_Time')

            if day_df.empty:
                continue

            # Day header
            ws.merge_cells(f'A{row}:I{row}')
            cell = ws[f'A{row}']
            cell.value = day.upper()
            cell.fill = self.day_fill
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center')
            row += 1

            # Column headers
            headers = ['Time', 'Course', 'Program', 'Section', 'Instructor', 'Room', 'Building', 'Type', 'Duration']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            row += 1

            # Data rows
            for _, entry in day_df.iterrows():
                ws.cell(row=row, column=1, value=f"{entry['Start_Time']} - {entry['End_Time']}")
                ws.cell(row=row, column=2, value=entry['Course_Name'])
                ws.cell(row=row, column=3, value=entry.get('Program', 'Unknown'))
                ws.cell(row=row, column=4, value=entry['Section'])
                ws.cell(row=row, column=5, value=entry['Instructor'])
                ws.cell(row=row, column=6, value=entry['Room'])
                ws.cell(row=row, column=7, value=entry.get('Building', 'N/A'))
                ws.cell(row=row, column=8, value=entry['Room_Type'])
                ws.cell(row=row, column=9, value=f"{entry['Duration_Minutes']} min")

                # Apply borders
                for col in range(1, 10):
                    ws.cell(row=row, column=col).border = self.border

                row += 1

            row += 1  # Empty row between days

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                # Skip merged cells
                if isinstance(cell, type(cell)) and hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter

                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

            if column_letter:
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(output_path)
        return output_path

    def _export_by_program(self, df: pd.DataFrame, output_path: str) -> str:
        """Export separate sheet for each program."""

        wb = Workbook()

        # 'Program' field is populated from Section.name in load_timetable_data
        programs = df['Program'].dropna().unique()

        # Filter out empty/unknown programs
        valid_programs = [p for p in programs if p and str(p).strip() and str(p).lower() != 'unknown']

        if not valid_programs:
            # No valid programs found - create an info sheet instead of empty workbook
            ws = wb.active
            ws.title = "No Programs Found"
            ws['A1'] = "No programs found in the timetable data."
            ws['A2'] = "Please ensure your dataset includes program information in the Section data."
            ws['A4'] = "Note: Program data comes from the 'program' column in your course dataset."
            wb.save(output_path)
            return output_path

        # Remove default sheet since we'll create program-specific sheets
        wb.remove(wb.active)

        for program in sorted(valid_programs):
            program_df = df[df['Program'] == program]

            # Create sheet with sanitized name
            safe_name = str(program).replace('/', '_').replace('\\', '_').replace(':', '-')[:31]
            ws = wb.create_sheet(title=safe_name)

            self._write_timetable_to_sheet(ws, program_df, f"Timetable for {program}")

        wb.save(output_path)
        return output_path

    def _export_free_slots(self, df: pd.DataFrame, output_path: str, timetable_id: int) -> str:
        """Export all unallocated/free time slots."""

        wb = Workbook()
        ws = wb.active
        ws.title = "Free Slots"

        try:
            # 1. Get configuration for time range
            # Assuming institution_id=1 for now as per other code
            config = self.db.query(ConstraintConfig).filter(
                ConstraintConfig.institution_id == 1,
                ConstraintConfig.is_active == True
            ).first()

            # Use config times or sensible defaults
            start_time_str = config.start_time if config and config.start_time else "08:00"
            end_time_str = config.end_time if config and config.end_time else "17:00"

            # 2. Get all rooms
            rooms = self.db.query(Room).filter(
                Room.institution_id == 1,
                Room.is_available == True,
                Room.is_deleted == False
            ).all()

            if not rooms:
                # No rooms available - provide helpful error message
                ws['A1'] = "Cannot generate free slots report."
                ws['A2'] = "No available rooms found in the database."
                ws['A4'] = "Please ensure rooms have been uploaded and are marked as available."
                ws['A5'] = "Rooms must have: is_available=True and is_deleted=False"
                wb.save(output_path)
                return output_path

            # 3. Generate all possible slots (30 min increments)
            from classsync_core.utils import parse_time, time_to_minutes

            start_min = time_to_minutes(parse_time(start_time_str))
            end_min = time_to_minutes(parse_time(end_time_str))

            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

            # 4. Identify occupied slots from timetable data
            # Map: (Room, Day, Minute) -> Occupied
            occupied = set()

            if not df.empty:
                for _, entry in df.iterrows():
                    room_code = entry.get('Room', '')
                    day = entry.get('Weekday', '')
                    start_time = entry.get('Start_Time', '')
                    end_time = entry.get('End_Time', '')

                    if not room_code or not day or not start_time or not end_time:
                        continue

                    try:
                        s_min = time_to_minutes(parse_time(str(start_time)))
                        e_min = time_to_minutes(parse_time(str(end_time)))

                        for m in range(s_min, e_min, 30):  # Mark every 30 min block
                            occupied.add((room_code, day, m))
                    except (ValueError, AttributeError):
                        # Skip entries with invalid time format
                        continue

            # 5. Find free slots
            free_slots = []

            for room in rooms:
                for day in days:
                    for m in range(start_min, end_min, 30):
                        if (room.code, day, m) not in occupied:
                            # Convert minutes back to time string
                            h = m // 60
                            mn = m % 60
                            time_str = f"{h:02d}:{mn:02d}"

                            # Safely get room_type value
                            room_type_str = room.room_type.value if room.room_type else 'Unknown'

                            free_slots.append({
                                'Day': day,
                                'Time': time_str,
                                'Room': room.code,
                                'Building': room.building or 'N/A',
                                'Room_Type': room_type_str
                            })

            # 6. Create DataFrame for free slots
            free_df = pd.DataFrame(free_slots)

            if free_df.empty:
                ws['A1'] = "No free slots found."
                ws['A2'] = "All room-time combinations are occupied."
                wb.save(output_path)
                return output_path

            # 7. Write to Excel - Headers
            headers = ['Day', 'Time', 'Room', 'Building', 'Room Type']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border

            # 8. Sort and write rows
            day_order = {d: i for i, d in enumerate(days)}
            free_df['Day_Order'] = free_df['Day'].map(day_order)
            free_df = free_df.sort_values(['Day_Order', 'Time', 'Room'])

            row = 2
            for _, entry in free_df.iterrows():
                ws.cell(row=row, column=1, value=entry['Day'])
                ws.cell(row=row, column=2, value=entry['Time'])
                ws.cell(row=row, column=3, value=entry['Room'])
                ws.cell(row=row, column=4, value=entry['Building'])
                ws.cell(row=row, column=5, value=entry['Room_Type'])

                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = self.border

                row += 1

            # 9. Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    if isinstance(cell, type(cell)) and hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                if column_letter:
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

            wb.save(output_path)
            return output_path

        except Exception as e:
            # Provide a meaningful error message in the export file
            ws['A1'] = "Error generating free slots report."
            ws['A2'] = f"Error: {str(e)}"
            ws['A4'] = "Please check:"
            ws['A5'] = "1. Rooms are uploaded and available"
            ws['A6'] = "2. Constraint configuration exists"
            ws['A7'] = "3. Time format is valid (HH:MM)"
            wb.save(output_path)
            return output_path


    def _export_by_section(self, df: pd.DataFrame, output_path: str) -> str:
        """Export separate sheet for each section."""

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        sections = df['Section'].unique()

        for section in sorted(sections):
            section_df = df[df['Section'] == section]

            # Create sheet
            ws = wb.create_sheet(title=str(section)[:31])  # Excel sheet name limit

            self._write_timetable_to_sheet(ws, section_df, f"Timetable for {section}")

        wb.save(output_path)
        return output_path

    def _export_by_teacher(self, df: pd.DataFrame, output_path: str) -> str:
        """Export separate sheet for each teacher."""

        wb = Workbook()
        wb.remove(wb.active)

        teachers = df['Instructor'].unique()

        for teacher in sorted(teachers):
            teacher_df = df[df['Instructor'] == teacher]

            # Create sheet (sanitize name for Excel)
            safe_name = str(teacher).replace('/', '_').replace('\\', '_')[:31]
            ws = wb.create_sheet(title=safe_name)

            self._write_timetable_to_sheet(ws, teacher_df, f"Timetable for {teacher}")

        wb.save(output_path)
        return output_path

    def _export_by_room(self, df: pd.DataFrame, output_path: str) -> str:
        """Export separate sheet for each room."""

        wb = Workbook()
        wb.remove(wb.active)

        rooms = df['Room'].unique()

        for room in sorted(rooms):
            room_df = df[df['Room'] == room]

            # Create sheet
            safe_name = str(room).replace('/', '_')[:31]
            ws = wb.create_sheet(title=safe_name)

            self._write_timetable_to_sheet(ws, room_df, f"Timetable for {room}")

        wb.save(output_path)
        return output_path

    def _write_timetable_to_sheet(self, ws, df: pd.DataFrame, title: str):
        """Helper method to write timetable data to a worksheet."""

        # Title
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = title
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')

        row = 3
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

        for day in days:
            day_df = df[df['Weekday'] == day].sort_values('Start_Time')

            if day_df.empty:
                continue

            # Day header
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws[f'A{row}']
            cell.value = day
            cell.fill = self.day_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            row += 1

            # Column headers
            headers = ['Time', 'Course', 'Section', 'Room', 'Instructor', 'Duration']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.border
            row += 1

            # Data
            for _, entry in day_df.iterrows():
                ws.cell(row=row, column=1, value=f"{entry['Start_Time']}-{entry['End_Time']}")
                ws.cell(row=row, column=2, value=entry['Course_Name'])
                ws.cell(row=row, column=3, value=entry['Section'])
                ws.cell(row=row, column=4, value=entry['Room'])
                ws.cell(row=row, column=5, value=entry['Instructor'])
                ws.cell(row=row, column=6, value=f"{entry['Duration_Minutes']} min")

                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = self.border

                row += 1

            row += 1

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                if isinstance(cell, type(cell)) and hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter

                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

            if column_letter:
                ws.column_dimensions[column_letter].width = min(max_length + 2, 40)